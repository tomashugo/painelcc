from django.shortcuts import render
from django.db.models import Q
from .forms import ProducaoForm, AlvosAbertos
import datetime
from django.http import HttpResponse, HttpResponseRedirect
from django.contrib.auth.decorators import login_required
import string, random, os
from openpyxl import Workbook
from .models import AlvosAbertos, AlvosDespachados, Producao
from mm.models import Company, MeterHistory
from analysis.models import Inspection
from .forms import AlvosAbertosForm, ReceberAlvoForm
from crews.models import Employee
# Create your views here.

def index(request):
   message = request.GET.get('message')

   try:
       company_session = Company.objects.get(name=request.session['Company'])
   except KeyError:
       company_session = "Equatorial"

   if request.method == 'POST':
      form = ProducaoForm(request.POST)
      if form.is_valid():              
       
         post = form.save(commit=False)
         post.envio = datetime.datetime.now()
	 post.save()

         return HttpResponseRedirect('/producao/?message=sucesso')
   else:
      form = ProducaoForm()
     
   return render(request,'producao/formv2.0.html',{'message':message,'form':form,'company_session':company_session,})

@login_required(login_url='/admin/login/')
def form_receber_alvo(request):
   company_session = Company.objects.get(name=request.session['Company'])
   message = request.GET.get('message')

   if request.method == 'GET':
      form = ReceberAlvoForm()
      id = request.GET.get('id')

      alvo = Producao.objects.get(id=id) 

      #versao = request.GET.get('versao')

      versao = '2.0'

      if versao == '2.0':
         return render(request,'producao/form_receber_alvov2.0.html',{'form' : form,'alvo':alvo,'company_session':company_session,})
      else:
         return render(request,'producao/form_receber_alvo.html',{'form' : form,'alvo':alvo,})

   else:
      form = ReceberAlvoForm(request.POST)
      if form.is_valid():
         id = request.GET.get('id')
         prod = Producao.objects.get(id=id)
         prod.data_devolucao = form.cleaned_data['data_devolucao']
         prod.save()
      
         return HttpResponseRedirect('/analysis/producao/?message=sucesso')


@login_required(login_url='/admin/login/')
def form_alvos_abertos(request):
   company_session = Company.objects.get(name=request.session['Company'])

   if request.method == 'GET':
      form = AlvosAbertosForm()
      id = request.GET.get('id')

      alvo = AlvosAbertos.objects.get(id=id) 

      #versao = request.GET.get('versao')

      versao = '2.0'

      if versao == '2.0':
         return render(request,'producao/form_alvos_abertosv2.0.html',{'alvo':alvo,'form' : form,'company_session':company_session,})
      else:
         return render(request,'producao/form_alvos_abertos.html',{'alvo':alvo,'form' : form,})

   else:
      form = AlvosAbertosForm(request.POST)
      if form.is_valid():
         id = request.GET.get('id')
         alvo = AlvosDespachados()
         alvo.inspetor = form.cleaned_data['inspetor']
         alvo.data_despacho = form.cleaned_data['data_despacho']
         alvo.alvo_aberto = AlvosAbertos.objects.get(id=id)
         alvo.save()
         return HttpResponseRedirect('/producao/alvos_abertos/?message=sucesso')
      return HttpResponse("Algo deu errado")


@login_required(login_url='/admin/login/')
def alvos_abertos(request):
   company_session = Company.objects.get(name=request.session['Company'])
   chars = string.letters + string.digits
   nome_arquivo = ''.join((random.choice(chars)) for x in range(20))

   class Output(object):
      pass

   class Periodo(object):
      pass

   outputs = []

   alarms = AlvosAbertos.objects.all()

   message = request.GET.get('message')
   #versao = request.GET.get('versao')
   versao = '2.0'

   for alarm in alarms:
      output = Output()

      output.id = alarm.id
      output.data_geracao = alarm.data_geracao
      output.consumer = alarm.consumer
      output.observacao = alarm.observacao
      output.ns = alarm.ns

      hoje = datetime.datetime.now()

      meter = MeterHistory.objects.all().filter(consumer=alarm.consumer).filter(since__lte=hoje).filter(until__gte=hoje)

      try:
         output.serial = meter.first().meter.serial
      except AttributeError:
         pass

      despachados = AlvosDespachados.objects.all().filter(alvo_aberto = alarm)
      executados = Producao.objects.all().filter(data_realizacao__gte=alarm.data_geracao).filter(instalacao=output.consumer.installation)
      
      output.executado = False

      if len(executados) > 0:
         output.executado = True

      if len(despachados) == 0:
         outputs.append(output)

   wb = Workbook()
   ws = wb.active

   ws.append(['Ns','Data Geracao','Instalacao','Medidor','Cliente','Localidade','Regional','Observacao'])

   for output in outputs:
      try:
         serial = output.serial
      except AttributeError:
         serial = ""

      ws.append([output.ns,output.data_geracao,output.consumer.installation,serial,output.consumer.name,output.consumer.city,output.consumer.region.name,output.observacao])

   xlsx = os.path.join("/home/tomash/painelcc/tmp",nome_arquivo+".xlsx")
   wb.save(xlsx)
 
   if versao == '2.0':
      return render(request,'producao/alvos_abertosv2.0.html',{'message':message,'company_session':company_session,'outputs': outputs,'arquivo': nome_arquivo,})
   else:
      return render(request,'producao/alvos_abertos.html',{'message':message,'company_session':company_session,'outputs': outputs,'arquivo': nome_arquivo,})

@login_required(login_url='/admin/login/')
def alvos_despachados(request):
   company_session = Company.objects.get(name=request.session['Company'])
   chars = string.letters + string.digits
   nome_arquivo = ''.join((random.choice(chars)) for x in range(20))

   class Output(object):
      pass

   class Periodo(object):
      pass

   outputs = []

   
   if request.method == 'GET':
       alarms = AlvosDespachados.objects.all().order_by('-alvo_aberto','-data_despacho')[:100]
   elif request.method == 'POST':
       filtrar_inspetor = request.POST.get('filtrar_inspetor')
       filtrar_prioridade = request.POST.get('filtrar_prioridade')
       inspetor = request.POST.get('inspetor')
       
       if filtrar_inspetor:
           alarms = AlvosDespachados.objects.all().filter(inspetor__id=inspetor)
       else:
           alarms = AlvosDespachados.objects.all().order_by('-alvo_aberto','-data_despacho')[:100]
           

   id_anterior = 0

   wb = Workbook()
   ws = wb.active

   ws.append(['Ns','Data Geracao','Instalacao','Cliente','Localidade','Regional','Observacao','Data Despacho','Inspetor'])

   tecnicos = []
   
   lista_tecnicos = Employee.objects.all()
   
   for i in lista_tecnicos:
       output = Output()
       
       output.id = i.id
       output.nome = i.name
       
       tecnicos.append(output)

   for alarm in alarms:
      output = Output()

      if id_anterior != alarm.alvo_aberto.id:
         output.data_geracao = alarm.alvo_aberto.data_geracao
         output.observacao = alarm.alvo_aberto.observacao
         output.ns = alarm.alvo_aberto.ns  
         output.installation = alarm.alvo_aberto.consumer.installation
         output.consumer_id = alarm.alvo_aberto.consumer.id
         output.name = alarm.alvo_aberto.consumer.name
         output.region = alarm.alvo_aberto.consumer.region.name     
         output.localidade = alarm.alvo_aberto.consumer.city
      else:
         output.consumer_id = ""
         output.localidade = ""
         output.installation = ""
         output.name = ""
         output.data_geracao = ""
         output.observacao = ""
         output.ns = ""
         output.region = ""


      id_anterior = alarm.alvo_aberto.id

      output.data_despacho = alarm.data_despacho
      output.id = alarm.alvo_aberto.id
      output.inspetor = alarm.inspetor

      executado = Inspection.objects.all().filter(ns=alarm.alvo_aberto.ns)
              
      if len(executado) == 0:               
         ws.append([alarm.alvo_aberto.ns,alarm.alvo_aberto.data_geracao,alarm.alvo_aberto.consumer.installation,alarm.alvo_aberto.consumer.name,alarm.alvo_aberto.consumer.city,alarm.alvo_aberto.consumer.region.name,alarm.alvo_aberto.observacao,alarm.data_despacho,unicode(alarm.inspetor)])

         outputs.append(output)


   xlsx = os.path.join("/home/tomash/painelcc/tmp",nome_arquivo+".xlsx")
   wb.save(xlsx)

   #versao = request.GET.get('versao')
   versao = '2.0'

   if versao == '2.0':
      return render(request,'producao/alvos_despachadosv2.0.html',{'company_session':company_session,'outputs': outputs,'arquivo': nome_arquivo,'tecnicos':tecnicos,})
   else:
      return render(request,'producao/alvos_despachados.html',{'company_session':company_session,'outputs': outputs,'arquivo': nome_arquivo,})


@login_required(login_url='/admin/login/')
def stats(request):
   company_session = Company.objects.get(name=request.session['Company'])

   class Output(object):
      pass

   class Periodo(object):
      pass

   outputs = []

   alarms = AlvosDespachados.objects.all().order_by('-alvo_aberto','-data_despacho')

   hoje = datetime.datetime.now()

   tempo_medio = dict()
   inspetor_id = dict()

   for i in Employee.objects.all():
      tempo_medio[i] = []
      inspetor_id[i.name] = i.id

   for alarm in alarms:
      diff = hoje.date()-alarm.data_despacho
      inspecoes = Inspection.objects.all().filter(ns = alarm.alvo_aberto.ns)

      if len(inspecoes) == 0:
         tempo_medio[alarm.inspetor].append(diff.days)

   a = ""

   outputs = []

   for keyvalue in tempo_medio.items():
      if len(tempo_medio[keyvalue[0]]) != 0:
         media = str(sum(tempo_medio[keyvalue[0]])/len(tempo_medio[keyvalue[0]])) + " dias"
         quantidade = str(len(tempo_medio[keyvalue[0]])) + " alvos"
         maior = str(max(tempo_medio[keyvalue[0]])) + " dias"

         output = Output()
         output.inspetor = str(keyvalue[0])
         output.id = inspetor_id[output.inspetor]
         output.media = media
         output.quantidade = quantidade
         output.maior = maior
      
         outputs.append(output)

   #return HttpResponse(a)
   
   #versao = request.GET.get('versao')
   versao = '2.0'

   if versao == '2.0':
       return render(request,'producao/statsv2.0.html',{'company_session':company_session,'outputs': outputs,})
   else:
       return render(request,'producao/stats.html',{'company_session':company_session,'outputs': outputs,})

@login_required(login_url='/admin/login/')
def alvos_nao_baixados(request):
   company_session = Company.objects.get(name=request.session['Company'])
   chars = string.letters + string.digits
   nome_arquivo = ''.join((random.choice(chars)) for x in range(20))

   class Output(object):
      pass

   class Periodo(object):
      pass

   outputs = []

   producoes = Producao.objects.all().filter(tipo_servico__id = 14)

   for p in producoes:
      output = Output()
    
      delta = datetime.timedelta(days=10)
      alvo_aberto = AlvosAbertos.objects.all().filter(Q(data_geracao__lte = p.data_realizacao) | Q(data_geracao__lte = p.data_realizacao + delta)).filter(consumer__installation = p.instalacao) 

      if len(alvo_aberto) > 0:          
         existe_alvo = 0
         
         for i in alvo_aberto:
             inspection = Inspection.objects.all().filter(ns = i.ns)
             
             if len(inspection) != 0:
                 existe_alvo = existe_alvo + 1
             else:
                 tmp = i

         if existe_alvo == 0:
            output.data_geracao = i.data_geracao
            output.consumer = i.consumer
            output.observacao = i.observacao
            output.ns = i.ns
            output.inspetor = p.tecnico
            output.data_realizacao = p.data_realizacao
            output.diferenca = datetime.datetime.now().date() - output.data_realizacao

            outputs.append(output)

   wb = Workbook()
   ws = wb.active

   ws.append(['Ns','Data Geracao','Instalacao','Cliente','Observacao','Data Realizacao','Inspetor','Diferenca'])

   for output in outputs:
      ws.append([output.ns,output.data_geracao,output.consumer.installation,output.consumer.name,output.observacao,output.data_realizacao,unicode(output.inspetor),output.diferenca])

   xlsx = os.path.join("/home/tomash/painelcc/tmp",nome_arquivo+".xlsx")
   wb.save(xlsx)

   #versao = request.GET.get('versao')

   versao = '2.0'

   if versao == '2.0':
      return render(request,'producao/alvos_nao_baixadosv2.0.html',{'company_session':company_session,'outputs': outputs,'arquivo': nome_arquivo,})
   else:
      return render(request,'producao/alvos_nao_baixados.html',{'company_session':company_session,'outputs': outputs,'arquivo': nome_arquivo,})
