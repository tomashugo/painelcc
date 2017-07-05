#!/usr/local/bin/python
#
# -*- coding: utf-8 -*-

from django.shortcuts import render
from django.http import HttpResponse, HttpResponseRedirect
from mm.models import Consumer, Mm, Quantity, MeterHistory, History, Change, Company
from datetime import datetime
from analysis.models import Billing, Inspection, RelatorioAlteracoesMedidor, RelatorioCorrenteZerada, RelatorioTensaoZerada, RelatorioMMVersusConsumo, RelatorioQuedaDeConsumo, RelatorioFraudeNaoIncrementada
from django.db import models
import random, string, os
from .forms import FormBillingMM, TheForm
from openpyxl import Workbook
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from datetime import timedelta
from django import forms
from django.contrib.auth.models import User
from producao.models import Producao

# Create your views here.
def dias_mes(inicio):
   month = inicio.month

   dias = []

   while inicio.month == month:
      dias.append(inicio)
      inicio = inicio + timedelta(days=1)

   return dias

@login_required(login_url='/admin/login/')
def produtividade(request):
   company_session = Company.objects.get(name=request.session['Company'])

   usernames = User.objects.all().distinct('username')

   usuarios = []

   for u in usernames:
      usuarios.append(u.username)

   class Output(object):
      pass

   report = []

   dia = request.GET.get('mes')

   dias = dias_mes( datetime.strptime(dia,"%d/%m/%Y") )

   dicionario = dict()
   dicionario_mm_versus_consumo = dict()
   dicionario_queda_de_consumo = dict()
   dicionario_alteracoes_medidor = dict()
   dicionario_corrente_zerada = dict()
   dicionario_tensao_zerada = dict()

   justificados_mm_versus_consumo = RelatorioMMVersusConsumo.objects.all().filter(justificado = True).filter(data_hora__month = dias[0].month).filter(company=company_session)
   justificados_queda_de_consumo = RelatorioQuedaDeConsumo.objects.all().filter(justificado = True).filter(data_hora__month = dias[0].month).filter(company=company_session)
   justificados_alteracoes_medidor = RelatorioAlteracoesMedidor.objects.all().filter(justificado = True).filter(data_hora__month = dias[0].month).filter(company=company_session)
   justificados_corrente_zerada = RelatorioCorrenteZerada.objects.all().filter(justificado = True).filter(data_hora__month = dias[0].month).filter(~Q(justificativa__contains = '!BATCH')).filter(company=company_session)
   justificados_tensao_zerada = RelatorioTensaoZerada.objects.all().filter(justificado = True).filter(data_hora__month = dias[0].month).filter(~Q(justificativa__contains = '!BATCH')).filter(company=company_session)


   for u in usuarios:
      dicionario[u] = dict()
      dicionario_mm_versus_consumo[u] = dict()
      dicionario_queda_de_consumo[u] = dict()
      dicionario_alteracoes_medidor[u] = dict()
      dicionario_corrente_zerada[u] = dict()
      dicionario_tensao_zerada[u] = dict()

      for d in dias:
         dicionario[u][d.strftime("%d/%m/%Y")] = 0
         dicionario_mm_versus_consumo[u][d.strftime("%d/%m/%Y")] = 0
         dicionario_queda_de_consumo[u][d.strftime("%d/%m/%Y")] = 0
         dicionario_alteracoes_medidor[u][d.strftime("%d/%m/%Y")] = 0
         dicionario_corrente_zerada[u][d.strftime("%d/%m/%Y")] = 0
         dicionario_tensao_zerada[u][d.strftime("%d/%m/%Y")] = 0

   for i in justificados_mm_versus_consumo:
      dicionario[i.user][i.data_hora.strftime("%d/%m/%Y")] = dicionario[i.user][i.data_hora.strftime("%d/%m/%Y")] + 1
      dicionario_mm_versus_consumo[i.user][i.data_hora.strftime("%d/%m/%Y")] = dicionario_mm_versus_consumo[i.user][i.data_hora.strftime("%d/%m/%Y")] + 1

   for i in justificados_queda_de_consumo:
      dicionario[i.user][i.data_hora.strftime("%d/%m/%Y")] = dicionario[i.user][i.data_hora.strftime("%d/%m/%Y")] + 1
      dicionario_queda_de_consumo[i.user][i.data_hora.strftime("%d/%m/%Y")] = dicionario_queda_de_consumo[i.user][i.data_hora.strftime("%d/%m/%Y")] + 1

   for i in justificados_alteracoes_medidor:
      dicionario[i.user][i.data_hora.strftime("%d/%m/%Y")] = dicionario[i.user][i.data_hora.strftime("%d/%m/%Y")] + 1
      dicionario_alteracoes_medidor[i.user][i.data_hora.strftime("%d/%m/%Y")] = dicionario_alteracoes_medidor[i.user][i.data_hora.strftime("%d/%m/%Y")] + 1

   for i in justificados_corrente_zerada:
      dicionario[i.user][i.data_hora.strftime("%d/%m/%Y")] = dicionario[i.user][i.data_hora.strftime("%d/%m/%Y")] + 1
      dicionario_corrente_zerada[i.user][i.data_hora.strftime("%d/%m/%Y")] = dicionario_corrente_zerada[i.user][i.data_hora.strftime("%d/%m/%Y")] + 1

   for i in justificados_tensao_zerada:  
      dicionario[i.user][i.data_hora.strftime("%d/%m/%Y")] = dicionario[i.user][i.data_hora.strftime("%d/%m/%Y")] + 1
      dicionario_tensao_zerada[i.user][i.data_hora.strftime("%d/%m/%Y")] = dicionario_tensao_zerada[i.user][i.data_hora.strftime("%d/%m/%Y")] + 1


   report = []
   report_mm_versus_consumo = []
   report_queda_de_consumo = []
   report_alteracoes_medidor = []
   report_corrente_zerada = []
   report_tensao_zerada = []

   for u in usuarios:
      linha = []
      linha_mm_versus_consumo = []
      linha_queda_de_consumo = []
      linha_alteracoes_medidor = []
      linha_corrente_zerada = []
      linha_tensao_zerada = []

      linha.append(u)
      linha_mm_versus_consumo.append(u)
      linha_queda_de_consumo.append(u)
      linha_alteracoes_medidor.append(u)
      linha_corrente_zerada.append(u)
      linha_tensao_zerada.append(u)


      for d in dias:
         linha.append(dicionario[u][d.strftime("%d/%m/%Y")])

         linha_mm_versus_consumo.append(dicionario_mm_versus_consumo[u][d.strftime("%d/%m/%Y")])
         linha_queda_de_consumo.append(dicionario_queda_de_consumo[u][d.strftime("%d/%m/%Y")])
	 linha_alteracoes_medidor.append(dicionario_alteracoes_medidor[u][d.strftime("%d/%m/%Y")])
	 linha_corrente_zerada.append(dicionario_corrente_zerada[u][d.strftime("%d/%m/%Y")])
	 linha_tensao_zerada.append(dicionario_tensao_zerada[u][d.strftime("%d/%m/%Y")])


      report.append(linha)
      report_mm_versus_consumo.append(linha_mm_versus_consumo)
      report_queda_de_consumo.append(linha_queda_de_consumo)
      report_alteracoes_medidor.append(linha_alteracoes_medidor)
      report_corrente_zerada.append(linha_corrente_zerada)
      report_tensao_zerada.append(linha_tensao_zerada)
      
   return render(request,'analysis/produtividade.html',{'company_session':company_session,'report':report, 'report_mm_versus_consumo': report_mm_versus_consumo, 'report_queda_de_consumo': report_queda_de_consumo,'report_alteracoes_medidor': report_alteracoes_medidor,'report_corrente_zerada':report_corrente_zerada,'report_tensao_zerada':report_tensao_zerada,'dias':dias,'usuarios':usuarios,'dicionario':dicionario,})

@login_required(login_url='/admin/login')
def ver_produtividade(request,analista,diames):
   company_session = Company.objects.get(name=request.session['Company'])

   dia = diames[0:2]
   mes = diames[2:4]
   ano = diames[4:8]

   mmversusconsumo = RelatorioMMVersusConsumo.objects.all().filter(justificado = True).filter(user = analista).filter(data_hora__day = dia).filter(data_hora__month = mes).filter(data_hora__year = ano).filter(company=company_session)
   quedadeconsumo = RelatorioQuedaDeConsumo.objects.all().filter(justificado = True).filter(user = analista).filter(data_hora__day = dia).filter(data_hora__month = mes).filter(data_hora__year = ano).filter(company=company_session)
   alteracoesmedidor = RelatorioAlteracoesMedidor.objects.all().filter(justificado = True).filter(user = analista).filter(data_hora__day = dia).filter(data_hora__month = mes).filter(data_hora__year = ano).filter(company=company_session)
   correntezerada = RelatorioCorrenteZerada.objects.all().filter(justificado = True).filter(user = analista).filter(data_hora__day = dia).filter(data_hora__month = mes).filter(data_hora__year = ano).filter(~Q(justificativa__contains = '!BATCH')).filter(company=company_session)
   tensaozerada = RelatorioTensaoZerada.objects.all().filter(justificado = True).filter(user = analista).filter(data_hora__day = dia).filter(data_hora__month = mes).filter(data_hora__year = ano).filter(~Q(justificativa__contains = '!BATCH')).filter(company=company_session)

   return render(request,'analysis/ver_produtividade.html',{'company_session':company_session,'mmversusconsumo':mmversusconsumo,'quedadeconsumo':quedadeconsumo,'alteracoesmedidor':alteracoesmedidor,'correntezerada':correntezerada,'tensaozerada':tensaozerada,'analista':analista,})

@login_required(login_url='/admin/login/')
def index(request):
   return HttpResponse("Aula 01 de Django - 27/12/2016")

@login_required(login_url='/admin/login/')
def form_alteracoes_medidor(request):
   company_session = Company.objects.get(name=request.session['Company'])

   if request.method == 'GET':
      form = TheForm()
      id = request.GET.get('id')

      #versao = request.GET.get('versao')
      versao = '2.0'

      if versao == '2.0':
         try:
            report = RelatorioAlteracoesMedidor.objects.get(id=id,justificado=False)
            return render(request,'analysis/form_alteracoes_medidorv2.0.html',{'company_session':company_session,'report':report,'form' : form,})
         except RelatorioAlteracoesMedidor.DoesNotExist:
            return HttpResponse('Erro no Formulario<br><a href="/analysis/alteracoes_medidor/">Voltar</a>')
      else:
         try:
            report = RelatorioAlteracoesMedidor.objects.get(id=id,justificado=False)
            return render(request,'analysis/form_alteracoes_medidor.html',{'company_session':company_session,'report':report,'form' : form,})
         except RelatorioAlteracoesMedidor.DoesNotExist:
            return HttpResponse('Erro no Formulario<br><a href="/analysis/alteracoes_medidor/">Voltar</a>')

   if request.method == 'POST':
      form = TheForm(request.POST)
      if form.is_valid():
         id = request.GET.get('id')
         report = RelatorioAlteracoesMedidor.objects.get(id=id)
         report.justificativa = form.cleaned_data['justificativa']
         report.user = request.user.username
         report.data_hora = datetime.now()
         report.justificado = True
         report.save()

         data_alteracao = report.data_alteracao - timedelta(hours=3)

	 reports = RelatorioAlteracoesMedidor.objects.all().filter(consumer=report.consumer).filter(data_alteracao__day = data_alteracao.day).filter(data_alteracao__month = data_alteracao.month).filter(data_alteracao__year = data_alteracao.year)

         for report in reports:
            report.justificativa = form.cleaned_data['justificativa']
            report.user = request.user.username
            report.data_hora = datetime.now()
            report.justificado = True
            report.save()

         return HttpResponse('Justificado com sucesso<BR><a href="/analysis/alteracoes_medidor/">Voltar</a>')
      else:
         return HttpResponse('ALGO DEU ERRADO NO FORMULARIO, TENTE NOVAMENTE<br><a href="/analysis/alteracoes_medidor/>Voltar</a>')

@login_required(login_url='/admin/login/')
def form_corrente_zerada(request):
   company_session = Company.objects.get(name=request.session['Company'])

   if request.method == 'GET':
      form = TheForm()
      id = request.GET.get('id')
      try:
         report = RelatorioCorrenteZerada.objects.get(id=id,justificado=False)

         history = History.objects.all().filter(consumer=report.consumer).filter(date_hour=report.inicio).filter(channel_1_qty__quantity = 'Ia').filter(channel_2_qty__quantity = 'Ib').filter(channel_3_qty__quantity = 'Ic')
         mm = history.first().mm
         mm_path = mm.path.replace('&','!')

         #versao = request.GET.get('versao')
         versao = '2.0'
         
         if versao == '2.0':
            return render(request,'analysis/form_corrente_zeradav2.0.html',{'company_session':company_session,'report':report,'form' : form,'mm': mm,'mm_path':mm_path,})
         else:
            return render(request,'analysis/form_corrente_zerada.html',{'company_session':company_session,'report':report,'form' : form,'mm': mm,'mm_path':mm_path,})
      except RelatorioCorrenteZerada.DoesNotExist:
         return HttpResponse('Erro no Formulario<br><a href="/analysis/corrente_zerada/">Voltar</a>')

   if request.method == 'POST':
      form = TheForm(request.POST)
      if form.is_valid():
         id = request.GET.get('id')
         expira = request.POST.get('expira')
         batch = request.POST.get('batch')

         if expira not in ["agora","3M","6M","12M"]:
            return HttpResponse('ERRO! Uma data para expiracao deve ser escolhida<br><a href="/analysis/form_corrente_zerada/?id='+id+'">Voltar</a>')

         data_expira = datetime.now()
         
         if expira == '3M':
            data_expira = data_expira + timedelta(days=90)
         if expira == '6M':
            data_expira = data_expira + timedelta(days=180)
         if expira == '12M':
            data_expira = data_expira + timedelta(days=365)
         
         report = RelatorioCorrenteZerada.objects.get(id=id)
         report.justificativa = form.cleaned_data['justificativa']
         report.user = request.user.username
         report.data_hora = datetime.now()
         report.data_expira = data_expira
         report.justificado = True
         report.save()

         if batch:
  	    reports = RelatorioCorrenteZerada.objects.all().filter(consumer=report.consumer).filter(justificado=False)

            for report in reports:
               report.justificativa = form.cleaned_data['justificativa'] + " !BATCH!"
               report.user = request.user.username
               report.data_expira = data_expira
               report.data_hora = datetime.now()
               report.justificado = True
               report.save()

         return HttpResponseRedirect('/analysis/corrente_zerada/?message=sucesso')
      else:
         return HttpResponse('ALGO DEU ERRADO NO FORMULARIO, TENTE NOVAMENTE<br><a href="/analysis/corrente_zerada/>Voltar</a>')

@login_required(login_url='/admin/login/')
def form_tensao_zerada(request):
   company_session = Company.objects.get(name=request.session['Company'])
   
   if request.method == 'GET':
      form = TheForm()
      id = request.GET.get('id')
      try:
         report = RelatorioTensaoZerada.objects.get(id=id,justificado=False)
         history = History.objects.all().filter(consumer=report.consumer).filter(date_hour=report.inicio).filter(channel_1_qty__quantity = 'Va').filter(channel_2_qty__quantity = 'Vb').filter(channel_3_qty__quantity = 'Vc')
         mm = history.first().mm
         mm_path = mm.path.replace('&','!')

         #versao = request.GET.get('versao')
         versao = '2.0' 

         if versao == '2.0':
            return render(request,'analysis/form_tensao_zeradav2.0.html',{'company_session':company_session,'report':report,'form' : form,'mm' : mm,'mm_path': mm_path,})
         else:
            return render(request,'analysis/form_tensao_zerada.html',{'company_session':company_session,'report':report,'form' : form,'mm' : mm,'mm_path': mm_path,})
      except RelatorioTensaoZerada.DoesNotExist:
         return HttpResponse('Erro no Formulario<br><a href="/analysis/tensao_zerada/">Voltar</a>')

   if request.method == 'POST':
      form = TheForm(request.POST)
      if form.is_valid():
         id = request.GET.get('id')
         expira = request.POST.get('expira')
         batch = request.POST.get('batch')

         if expira not in ["agora","3M","6M","12M"]:
            return HttpResponse('ERRO! Uma data para expiracao deve ser escolhida<br><a href="/analysis/form_tensao_zerada/?id='+id+'">Voltar</a>')

         data_expira = datetime.now()
         
         if expira == '3M':
            data_expira = data_expira + timedelta(days=90)
         if expira == '6M':
            data_expira = data_expira + timedelta(days=180)
         if expira == '12M':
            data_expira = data_expira + timedelta(days=365)
         
         report = RelatorioTensaoZerada.objects.get(id=id)
         report.justificativa = form.cleaned_data['justificativa']
         report.user = request.user.username
         report.data_hora = datetime.now()
         report.data_expira = data_expira
         report.justificado = True
         report.save()

         if batch:
  	    reports = RelatorioTensaoZerada.objects.all().filter(consumer=report.consumer).filter(justificado=False)

            for report in reports:
               report.justificativa = form.cleaned_data['justificativa'] + " !BATCH!"
               report.user = request.user.username
               report.data_expira = data_expira
               report.data_hora = datetime.now()
               report.justificado = True
               report.save()

         return HttpResponseRedirect('/analysis/tensao_zerada/?message=sucesso')
      else:
         return HttpResponse('ALGO DEU ERRADO NO FORMULARIO, TENTE NOVAMENTE<br><a href="/analysis/tensao_zerada/>Voltar</a>')


@login_required(login_url='/admin/login/')
def alteracoes_medidor(request):
   company_session = Company.objects.get(name=request.session['Company'])

   chars = string.letters + string.digits
   nome_arquivo = ''.join((random.choice(chars)) for x in range(20))

   reports = RelatorioAlteracoesMedidor.objects.all().filter(justificado = False).filter(company=company_session)

   class Output(object):
      pass

   outputs = []

   for report in reports:
      output = Output()

      output.consumer = report.consumer
      output.id = report.id
      output.instalacao = report.consumer.installation
      output.nome = report.consumer.name
      output.data = report.data_alteracao
      output.leitor = report.leitor
      output.code = report.code
      output.observacao = report.observacao

      outputs.append(output)

   wb = Workbook()
   ws = wb.active

   ws.append(['Instalacao','Nome','Data','Leitor','Code','Observacao','Inspecao','Leitura','Acao'])

   for output in outputs:
      ws.append([output.instalacao,output.nome,output.data,output.leitor,output.code,output.observacao])

   xlsx = os.path.join("/home/tomash/painelcc/tmp",nome_arquivo+".xlsx")
   wb.save(xlsx)

   #versao = request.GET.get('versao')

   versao = '2.0'

   if versao == '2.0':
      return render(request,'analysis/alteracoes_medidorv2.0.html',{'company_session':company_session,'outputs': outputs,'arquivo':nome_arquivo,})
   else:
      return render(request,'analysis/alteracoes_medidor.html',{'company_session':company_session,'outputs': outputs,'arquivo':nome_arquivo,})

@login_required(login_url='/admin/login/')
def form_queda_consumo(request):
   company_session = Company.objects.get(name=request.session['Company'])

   if request.method == 'GET':
      form = TheForm()
      id = request.GET.get('id')
      
      #versao = request.GET.get('versao')
 
      versao = '2.0'

      if versao == '2.0':
         try:
            report = RelatorioQuedaDeConsumo.objects.get(id=id,justificado=False)
            return render(request,'analysis/form_queda_consumov2.0.html',{'report':report,'form' : form,'company_session':company_session,})
         except RelatorioQuedaDeConsumo.DoesNotExist:
            return HttpResponse('Erro no Formulario<br><a href="/analysis/queda_consumo/">Voltar</a>')
      else:
         try:
            report = RelatorioQuedaDeConsumo.objects.get(id=id,justificado=False)
            return render(request,'analysis/form_queda_consumo.html',{'report':report,'form' : form,'company_session':company_session,})
         except RelatorioQuedaDeConsumo.DoesNotExist:
            return HttpResponse('Erro no Formulario<br><a href="/analysis/queda_consumo/">Voltar</a>')

   if request.method == 'POST':
      form = TheForm(request.POST)
      if form.is_valid():
         id = request.GET.get('id')
         expira = request.POST.get('expira')

         if expira not in ["agora","3M","6M","12M"]:
            return HttpResponse('ERRO! Uma data para expiracao deve ser escolhida<br><a href="/analysis/form_queda_consumo/?id='+id+'">Voltar</a>')

         data_expira = datetime.now()
         
         if expira == '3M':
            data_expira = data_expira + timedelta(days=90)
         if expira == '6M':
            data_expira = data_expira + timedelta(days=180)
         if expira == '12M':
            data_expira = data_expira + timedelta(days=365)
     
         if request.POST.get('alvo_gerado') is None:
            alvo_gerado = False
         else: 
            alvo_gerado = True
 
         report = RelatorioQuedaDeConsumo.objects.get(id=id)
         report.justificativa = form.cleaned_data['justificativa']
         report.user = request.user.username
         report.data_hora = datetime.now()
         report.data_expira = data_expira
         report.justificado = True
         report.alvo_gerado = alvo_gerado
         report.save()

         return HttpResponseRedirect('/analysis/queda_consumo/?message=sucesso')
      else:
         return HttpResponse('ALGO DEU ERRADO NO FORMULARIO, TENTE NOVAMENTE<br><a href="/analysis/queda_consumo/>Voltar</a>')
  
   return render(request,'analysis/form_queda_consumo.html',{'company_session':company_session,'outputs': outputs,'arquivo': nome_arquivo,})

@login_required(login_url='/admin/login/')
def queda_consumo(request):
   company_session = Company.objects.get(name=request.session['Company'])
   chars = string.letters + string.digits
   nome_arquivo = ''.join((random.choice(chars)) for x in range(20))

   message = request.GET.get('message')

   class Output(object):
      pass

   class Periodo(object):
      pass

   outputs = []

   alarms = RelatorioQuedaDeConsumo.objects.all().filter(justificado = False).order_by('consumer','-referencia').filter(company=company_session)

   for alarm in alarms:
      output = Output()

      output.id = alarm.id
      output.consumer = alarm.consumer
      output.referencia = alarm.referencia
      output.media_consumo = alarm.media_consumo
      output.consumo_referencia = alarm.consumo_referencia
      try:
         output.queda_percentual = (alarm.consumo_referencia-alarm.media_consumo)/alarm.media_consumo*100
      except ZeroDivisionError:
         output.queda_percentual = "100"
      outputs.append(output)

   wb = Workbook()
   ws = wb.active

   ws.append(['Instalacao','Nome','Referencia','Consumo Referencia','Media de Consumo','Queda Percentual'])

   for output in outputs:
      ws.append([output.consumer.installation,output.consumer.name,output.referencia,output.consumo_referencia,output.media_consumo,output.queda_percentual])

   xlsx = os.path.join("/home/tomash/painelcc/tmp",nome_arquivo+".xlsx")
   wb.save(xlsx)

   #versao = request.GET.get('versao')

   versao = '2.0'

   if versao == '2.0':
      return render(request,'analysis/queda_consumov2.0.html',{'message':message,'company_session':company_session,'outputs': outputs,'arquivo': nome_arquivo,})
   else:
      return render(request,'analysis/queda_consumo.html',{'message':message,'company_session':company_session,'outputs': outputs,'arquivo': nome_arquivo,})


@login_required(login_url='/admin/login/')
def producao(request):
   company_session = Company.objects.get(name=request.session['Company'])
   chars = string.letters + string.digits
   nome_arquivo = ''.join((random.choice(chars)) for x in range(20))

   class Output(object):
      pass

   class Periodo(object):
      pass

   outputs = []

   producao = Producao.objects.all().order_by('-id')

   message = request.GET.get('message')

   for p in producao:
      output = Output()

      output.id = p.id
      output.envio = p.envio
      output.tipo_servico = p.tipo_servico
      output.tecnico = p.tecnico
      output.instalacao = p.instalacao
      output.regional = p.regional
      output.codigo = p.codigo
      output.familia_codigo = p.familia_codigo
      output.data_realizacao = p.data_realizacao
      output.observacao = p.observacao
      output.data_devolucao = p.data_devolucao

      outputs.append(output)

   wb = Workbook()
   ws = wb.active

   ws.append(['Envio','Tipo Servico','Tecnico','Instalacao','Regional','Codigo','Familia Codigo','Data Realizacao','Observacao','Data Devolucao'])

   for output in outputs:
      ws.append([output.envio,unicode(output.tipo_servico),unicode(output.tecnico),output.instalacao,unicode(output.regional),output.codigo,output.familia_codigo,output.data_realizacao,unicode(output.observacao),output.data_devolucao])

   xlsx = os.path.join("/home/tomash/painelcc/tmp",nome_arquivo+".xlsx")
   wb.save(xlsx)

   #versao = request.GET.get('versao')
 
   versao = '2.0'

   if versao == '2.0':
      return render(request,'analysis/producaov2.0.html',{'message':message,'company_session':company_session,'outputs': outputs,'arquivo': nome_arquivo,})
   else:
      return render(request,'analysis/producao.html',{'message':message,'company_session':company_session,'outputs': outputs,'arquivo': nome_arquivo,})


@login_required(login_url='/admin/login/')
def fraude_nao_incrementada(request):
   company_session = Company.objects.get(name=request.session['Company'])
   chars = string.letters + string.digits
   nome_arquivo = ''.join((random.choice(chars)) for x in range(20))

   class Output(object):
      pass

   class Periodo(object):
      pass

   outputs = []

   message = request.GET.get('message')

   alarms = RelatorioFraudeNaoIncrementada.objects.all().filter(justificado = False).order_by('-mes_fraude').filter(company=company_session)

   for alarm in alarms:
      output = Output()

      output.id = alarm.id
      output.consumer = alarm.consumer
      output.mes_fraude = alarm.mes_fraude
      output.code_fraude = alarm.code_fraude
      output.faturamento_anterior = alarm.faturamento_anterior
      output.faturamento_posterior = alarm.faturamento_posterior

      outputs.append(output)

   wb = Workbook()
   ws = wb.active

   ws.append(['Instalacao','Nome','Mes Fraude','Code','Faturamento m-1','Faturamento m+1'])

   for output in outputs:
      ws.append([output.consumer.installation,output.consumer.name,output.mes_fraude,output.code_fraude,output.faturamento_anterior,output.faturamento_posterior])

   xlsx = os.path.join("/home/tomash/painelcc/tmp",nome_arquivo+".xlsx")
   wb.save(xlsx)

   #versao = request.GET.get('versao')

   versao = '2.0'

   if versao == '2.0':
      return render(request,'analysis/fraude_nao_incrementadav2.0.html',{'message':message,'company_session':company_session,'outputs': outputs,'arquivo': nome_arquivo,'quantidade':len(outputs),})
   else:
      return render(request,'analysis/fraude_nao_incrementada.html',{'message':message,'company_session':company_session,'outputs': outputs,'arquivo': nome_arquivo,'quantidade':len(outputs),})

@login_required(login_url='/admin/login/')
def form_fraude_nao_incrementada(request):
   company_session = Company.objects.get(name=request.session['Company'])

   if request.method == 'GET':
      form = TheForm()
      id = request.GET.get('id')

      #versao = request.GET.get('versao')

      versao = '2.0'
   
      if versao == '2.0':
         try:
            report = RelatorioFraudeNaoIncrementada.objects.get(id=id,justificado=False)
            return render(request,'analysis/form_fraude_nao_incrementadav2.0.html',{'report':report,'form' : form,'company_session':company_session,})
         except RelatorioQuedaDeConsumo.DoesNotExist:
            return HttpResponse('Erro no Formulario<br><a href="/analysis/fraude_nao_incrementada/">Voltar</a>')
      else:
         try:
            report = RelatorioFraudeNaoIncrementada.objects.get(id=id,justificado=False)
            return render(request,'analysis/form_fraude_nao_incrementada.html',{'report':report,'form' : form,})
         except RelatorioQuedaDeConsumo.DoesNotExist:
            return HttpResponse('Erro no Formulario<br><a href="/analysis/fraude_nao_incrementada/">Voltar</a>')

   if request.method == 'POST':
      form = TheForm(request.POST)
      if form.is_valid():
         id = request.GET.get('id')
         
         report = RelatorioFraudeNaoIncrementada.objects.get(id=id)
         report.justificativa = form.cleaned_data['justificativa']
         report.user = request.user.username
         report.data_hora = datetime.now()
         report.justificado = True
         report.save()

         return HttpResponseRedirect('/analysis/fraude_nao_incrementada/?message=sucesso')
      else:
         return HttpResponse('ALGO DEU ERRADO NO FORMULARIO, TENTE NOVAMENTE<br><a href="/analysis/relatorio_fraude_nao_incrementada/>Voltar</a>')
  
   return render(request,'analysis/form_fraude_nao_incrementada.html',{'company_session':company_session,'outputs': outputs,'arquivo': nome_arquivo,})



@login_required(login_url='/admin/login/')
def corrente_zerada(request):
   company_session = Company.objects.get(name=request.session['Company'])
   chars = string.letters + string.digits
   nome_arquivo = ''.join((random.choice(chars)) for x in range(20))

   class Output(object):
      pass

   class Periodo(object):
      pass

   outputs = []

   alarms = RelatorioCorrenteZerada.objects.all().filter(justificado = False).filter(company=company_session).distinct('consumer')

   recorrencia = dict()

   #versao = request.GET.get('versao')
   versao = '2.0'

   message = request.GET.get('message')

   for alarm in alarms:
      output = Output()

      recorrencia = RelatorioCorrenteZerada.objects.all().filter(justificado= False).filter(consumer=alarm.consumer)

      output.id = alarm.id
      output.consumer = alarm.consumer
      output.inicio = alarm.inicio
      output.fim = alarm.fim
      output.ia = alarm.ia
      output.ib = alarm.ib
      output.ic = alarm.ic
      output.link_mm = "?data_de="+(alarm.inicio-timedelta(hours=3)).strftime("%d/%m/%Y %H:%M")+"&data_ate="+(alarm.fim-timedelta(hours=3)).strftime("%d/%m/%Y %H:%M")
      output.recorrencia = len(recorrencia)

      outputs.append(output)

      anterior = alarm.consumer

   wb = Workbook()
   ws = wb.active

   ws.append(['Instalacao','Faturamento','Nome','Inicio','Fim','Recorrencia','Ia','Ib','Ic'])

   for output in outputs:
      ws.append([output.consumer.installation,output.consumer.revenue,output.consumer.name,output.inicio,output.fim,output.recorrencia,output.ia,output.ib,output.ic])

   xlsx = os.path.join("/home/tomash/painelcc/tmp",nome_arquivo+".xlsx")
   wb.save(xlsx)
   
   tem_recorrencia = True

   if versao == '2.0':
      return render(request,'analysis/corrente_zeradav2.0.html',{'message':message,'company_session':company_session,'outputs': outputs,'arquivo': nome_arquivo,'tem_recorrencia' : tem_recorrencia,})
   else:
      return render(request,'analysis/corrente_zerada.html',{'message':message,'company_session':company_session,'outputs': outputs,'arquivo': nome_arquivo,'tem_recorrencia' : tem_recorrencia,})

@login_required(login_url='/admin/login/')
def recorrencia_corrente_zerada(request,consumer_id):
   company_session = Company.objects.get(name=request.session['Company'])

   chars = string.letters + string.digits
   nome_arquivo = ''.join((random.choice(chars)) for x in range(20))

   class Output(object):
      pass

   class Periodo(object):
      pass

   outputs = []

   consumer = Consumer.objects.all().filter(id = consumer_id)

   alarms = RelatorioCorrenteZerada.objects.all().filter(justificado = False).filter(consumer = consumer).order_by('inicio')
      
   recorrencia = dict()

   if len(alarms) > 0:
      anterior = alarms.last().consumer

   for alarm in alarms:
      output = Output()

      try:
         recorrencia[alarm.consumer.installation] = recorrencia[alarm.consumer.installation] + 1
      except KeyError:
         recorrencia[alarm.consumer.installation] = 1
    
      output.id = alarm.id
      output.consumer = alarm.consumer
      output.inicio = alarm.inicio
      output.fim = alarm.fim
      output.ia = alarm.ia
      output.ib = alarm.ib
      output.ic = alarm.ic
      output.link_mm = "?data_de="+(alarm.inicio-timedelta(hours=3)).strftime("%d/%m/%Y %H:%M")+"&data_ate="+(alarm.fim-timedelta(hours=3)).strftime("%d/%m/%Y %H:%M")

      if alarm.consumer != anterior:
         output.recorrencia = recorrencia[alarm.consumer.installation]
         
      outputs.append(output)

      anterior = alarm.consumer

   tem_recorrencia = False

   #versao = request.GET.get('versao')

   versao = '2.0'

   if versao == '2.0':
      return render(request,'analysis/corrente_zeradav2.0.html',{'company_session':company_session,'outputs': outputs,'arquivo': nome_arquivo,'tem_recorrencia':tem_recorrencia,})
   else:
      return render(request,'analysis/corrente_zerada.html',{'company_session':company_session,'outputs': outputs,'arquivo': nome_arquivo,'tem_recorrencia':tem_recorrencia,})


@login_required(login_url='/admin/login/')
def recorrencia_tensao_zerada(request,consumer_id):
   company_session = Company.objects.get(name=request.session['Company'])

   chars = string.letters + string.digits
   nome_arquivo = ''.join((random.choice(chars)) for x in range(20))

   class Output(object):
      pass

   class Periodo(object):
      pass

   outputs = []

   consumer = Consumer.objects.all().filter(id = consumer_id)

   alarms = RelatorioTensaoZerada.objects.all().filter(justificado = False).filter(consumer = consumer).order_by('inicio')
      
   recorrencia = dict()

   if len(alarms) > 0:
      anterior = alarms.last().consumer

   for alarm in alarms:
      output = Output()

      try:
         recorrencia[alarm.consumer.installation] = recorrencia[alarm.consumer.installation] + 1
      except KeyError:
         recorrencia[alarm.consumer.installation] = 1
    
      output.id = alarm.id
      output.consumer = alarm.consumer
      output.inicio = alarm.inicio
      output.fim = alarm.fim
      output.va = alarm.va
      output.vb = alarm.vb
      output.vc = alarm.vc
      output.link_mm = "?data_de="+(alarm.inicio-timedelta(hours=3)).strftime("%d/%m/%Y %H:%M")+"&data_ate="+(alarm.fim-timedelta(hours=3)).strftime("%d/%m/%Y %H:%M")

      if alarm.consumer != anterior:
         output.recorrencia = recorrencia[alarm.consumer.installation]
         
      outputs.append(output)

      anterior = alarm.consumer
   
   tem_recorrencia = False

   #versao = request.GET.get('versao')

   versao = '2.0'

   if versao == '2.0':
      return render(request,'analysis/tensao_zeradav2.0.html',{'company_session':company_session,'outputs': outputs,'arquivo': nome_arquivo,'tem_recorrencia':tem_recorrencia,})
   else:
      return render(request,'analysis/tensao_zerada.html',{'company_session':company_session,'outputs': outputs,'arquivo': nome_arquivo,'tem_recorrencia':tem_recorrencia,})


@login_required(login_url='/admin/login/')
def tensao_zerada(request):
   company_session = Company.objects.get(name=request.session['Company'])
   chars = string.letters + string.digits
   nome_arquivo = ''.join((random.choice(chars)) for x in range(20))

   message = request.GET.get('message')

   class Output(object):
      pass

   class Periodo(object):
      pass

   outputs = []

   alarms = RelatorioTensaoZerada.objects.all().filter(justificado = False).filter(company=company_session).distinct('consumer')
      
   recorrencia = dict()

   for alarm in alarms:
      output = Output()

      recorrencia = RelatorioTensaoZerada.objects.all().filter(consumer=alarm.consumer).filter(justificado=False)
    
      output.consumer = alarm.consumer
      output.inicio = alarm.inicio
      output.fim = alarm.fim
      output.va = alarm.va
      output.vb = alarm.vb
      output.vc = alarm.vc
      output.link_mm = "?data_de="+(alarm.inicio-timedelta(hours=3)).strftime("%d/%m/%Y %H:%M")+"&data_ate="+(alarm.fim-timedelta(hours=3)).strftime("%d/%m/%Y %H:%M")

      output.recorrencia = len(recorrencia)

      outputs.append(output)


   wb = Workbook()
   ws = wb.active

   ws.append(['Instalacao','Nome','Inicio','Fim','Recorrencia','Va','Vb','Vc'])

   for output in outputs:
      ws.append([output.consumer.installation,output.consumer.name,output.inicio,output.fim,output.recorrencia,output.va,output.vb,output.vc])

   xlsx = os.path.join("/home/tomash/painelcc/tmp",nome_arquivo+".xlsx")
   wb.save(xlsx)

   tem_recorrencia = True

   #versao = request.GET.get('versao')

   versao = '2.0'

   if versao == '2.0':
      return render(request,'analysis/tensao_zeradav2.0.html',{'message':message,'company_session':company_session,'outputs': outputs,'arquivo': nome_arquivo,'tem_recorrencia':tem_recorrencia,})
   else:   
      return render(request,'analysis/tensao_zerada.html',{'message':message,'company_session':company_session,'outputs': outputs,'arquivo': nome_arquivo,'tem_recorrencia':tem_recorrencia,})


@login_required(login_url='/admin/login/')
def tc_saturado(request):
   company_session = Company.objects.get(name=request.session['Company'])
   chars = string.letters + string.digits
   nome_arquivo = ''.join((random.choice(chars)) for x in range(20))

   if request.method == 'POST':
      form = FormBillingMM(request.POST)
   else:
      form = FormBillingMM()

   if form.is_valid():
      mes = form.cleaned_data['referencia']
   else:
      mes = datetime.strptime("01/09/16","%d/%m/%y")

   consumers = Consumer.objects.all()

   ia = Quantity.objects.get(quantity = 'Ia')
   ib = Quantity.objects.get(quantity = 'Ib')
   ic = Quantity.objects.get(quantity = 'Ic')

   class Output(object):
      pass

   class Periodo(object):
      pass

   outputs = []

   mms = History.objects.all().filter(date_hour__month = mes.month).filter(date_hour__year = mes.year).filter(channel_1_qty = ia).filter(channel_2_qty = ib).filter(channel_3_qty = ic).filter(Q(channel_1_value__gte = 7) | Q(channel_2_value__gte = 7) | Q(channel_3_value__gte = 7)).filter(~Q(mm__serial__startswith = '05')).filter(~Q(mm__serial__startswith = '03')).distinct('mm')

   mm_analisadas = []

   for mm in mms:
      history = History.objects.all().filter(mm=mm.mm).filter(date_hour__month = mes.month).filter(date_hour__year = mes.year).order_by('date_hour')
      consumer = mm.consumer

      output = Output()
      output.instalacao = consumer.installation
      output.nome = consumer.name
      output.id = consumer.id

      zerado = False

      output.periodos = []

      for h in history:
         hora = h.date_hour

         if h.channel_1_value > 7.0 or h.channel_2_value > 7.0 or h.channel_3_value > 7.0:
            if zerado == False:
               zerado = True
               a = Periodo()
               a.hora_inicio = hora
               a.ia = h.channel_1_value
               a.ib = h.channel_2_value
               a.ic = h.channel_3_value
               output.periodos.append(a)
         else:
            if zerado == True:
               zerado = False
               output.periodos[-1].hora_fim = hora
   
               dif = hora-output.periodos[-1].hora_inicio

               if dif.seconds < 3600:
                  output.periodos.pop()


      outputs.append(output)
	
   wb = Workbook()
   ws = wb.active

   ws.append(['Mes Referencia',mes.strftime("%d/%m/%Y")])
   ws.append(['Instalacao','Cliente','Hora Inicio','Hora Fim','Ia','Ib','Ic'])

   for output in outputs:
      for periodo in output.periodos:
         try:
            ws.append([output.instalacao,output.nome,periodo.hora_inicio,periodo.hora_fim,periodo.ia,periodo.ib,periodo.ic])
            periodo.link_mm = "?data_de="+(periodo.hora_inicio-timedelta(hours=3)).strftime("%d/%m/%Y %H:%M")+"&data_ate="+(periodo.hora_fim-timedelta(hours=3)).strftime("%d/%m/%Y %H:%M")
         except AttributeError:
            ws.append([output.instalacao,output.nome,periodo.hora_inicio," ",periodo.ia,periodo.ib,periodo.ic])	  
            data_ate = periodo.hora_inicio + timedelta(days=30)
            periodo.link_mm = "?data_de="+(periodo.hora_inicio-timedelta(hours=3)).strftime("%d/%m/%Y %H:%M")+"&data_ate="+data_ate.strftime("%d/%m/%Y %H:%M")

   xlsx = os.path.join("/home/tomash/painelcc/tmp",nome_arquivo+".xlsx")
   wb.save(xlsx)

   return render(request,'analysis/tc_saturado.html',{'company_session': company_session,'outputs': outputs,'form': form,'arquivo': nome_arquivo,})
        
@login_required(login_url='/admin/login/')
def download_mm(request):
   arquivo = request.GET.get('arquivo')

   arquivo = arquivo.replace('!','&')

   publico = arquivo.split('/')
   publico = publico[len(publico)-1]

   response = HttpResponse(open(arquivo,'rb').read())
   response['Content-Type'] = 'mimetype/submimetype'
   response['Content-Disposition'] = 'attachment; filename='+publico

   return response

@login_required(login_url='/admin/login/')
def download(request):
   arquivo = request.GET.get('arquivo')

   file_path = os.path.join('/home/tomash/painelcc/tmp',arquivo)
   response = HttpResponse(open(file_path,'rb').read())
   response['Content-Type'] = 'mimetype/submimetype'
   response['Content-Disposition'] = 'attachment; filename=tmp.xlsx'

   return response

@login_required(login_url='/admin/login/')
def form_mm_versus_consumo(request):
   company_session = Company.objects.get(name=request.session['Company'])

   if request.method == 'GET':
      form = TheForm()
      id = request.GET.get('id')

      #versao = request.GET.get('versao')
 
      versao = '2.0'

      if versao == '2.0':
         try:
            report = RelatorioMMVersusConsumo.objects.get(id=id,justificado=False)
            return render(request,'analysis/form_mm_versus_consumov2.0.html',{'company_session':company_session,'report':report,'form' : form,})
         except RelatorioMMVersusConsumo.DoesNotExist:
            return HttpResponse('Erro no Formulario<br><a href="/analysis/mm_versus_consumo/">Voltar</a>')
      else:
         try:
            report = RelatorioMMVersusConsumo.objects.get(id=id,justificado=False)
            return render(request,'analysis/form_mm_versus_consumo.html',{'company_session':company_session,'report':report,'form' : form,})
         except RelatorioMMVersusConsumo.DoesNotExist:
            return HttpResponse('Erro no Formulario<br><a href="/analysis/mm_versus_consumo/">Voltar</a>')

   if request.method == 'POST':
      form = TheForm(request.POST)
      if form.is_valid():
         id = request.GET.get('id')
         report = RelatorioMMVersusConsumo.objects.get(id=id)
         report.justificativa = form.cleaned_data['justificativa']
         report.user = request.user.username
         report.data_hora = datetime.now()
         report.justificado = True
         report.save()
         return HttpResponse('Justificado com sucesso<BR><a href="/analysis/mm_versus_consumo/">Voltar</a>')
      else:
         return HttpResponse('ALGO DEU ERRADO NO FORMULARIO, TENTE NOVAMENTE<br><a href="/analysis/mm_versus_consumo/>Voltar</a>')

@login_required(login_url='/admin/login/')
def mm_versus_consumo(request):
   company_session = Company.objects.get(name=request.session['Company'])
   chars = string.letters + string.digits
   nome_arquivo = ''.join((random.choice(chars)) for x in range(20))

   clientes = []

   class Cliente(object):
      pass

   reports = RelatorioMMVersusConsumo.objects.all().filter(justificado=False).filter(company=company_session).order_by('mes_referencia')

   wb = Workbook()
   ws = wb.active

   ws.append(['Instalacao','Cliente','Fat','Medidor','Referencia','Faturado','Registrado MM','%','MM','Tipo Comparacao'])

   for report in reports:
      try: 
         dif = (report.consumo_faturado-report.consumo_mm)/report.consumo_mm*100
      except ZeroDivisionError:
         dif = 100
      
      try:
         report.dif = dif
         report.mm_path = report.mm.path.replace('&','!')
         ws.append([report.consumer.installation,report.consumer.name,report.consumer.revenue,report.meter.serial,report.mes_referencia,report.consumo_faturado,report.consumo_mm,dif,report.mm.archive,report.tipo_comparacao])
      except AttributeError:      
         pass

   xlsx = os.path.join("/home/tomash/painelcc/tmp",nome_arquivo+".xlsx")
   wb.save(xlsx)

   #versao = request.GET.get('versao')

   versao = '2.0'
   
   if versao == '2.0':
      return render(request,'analysis/mm_versus_consumov2.0.html',{'company_session':company_session,'reports': reports,'arquivo': nome_arquivo,})
   else:
      return render(request,'analysis/mm_versus_consumo.html',{'company_session':company_session,'reports': reports,'arquivo': nome_arquivo,})


